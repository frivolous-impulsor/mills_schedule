from openpyxl import load_workbook
import os
import csv
from datetime import datetime
import re
from enum import Enum

class Day(Enum):
    SUN = 0
    MON = 1
    TUE = 2
    WED = 3
    THU = 4
    FRI = 5
    SAT = 6


MAX_SLOT_IN_DAY: int = 10
MAX_DAY_IN_WEEK: int = 7

def writeCSV(filePath: str, matrix: list[list[int]]):
    for i, row in enumerate(matrix):
        matrix[i] = [list(Day)[i].name]+row 
    with open(filePath, 'w', newline='') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerows(matrix)

def getTimeDiff(initT: str, finalT: str)->int:
    if("am" in initT.lower() or "pm" in initT.lower() or "am" in finalT.lower() or "pm" in finalT.lower()):
        try: 
            initObj = datetime.strptime(initT, "%I:%M %p")
            finalObj = datetime.strptime(finalT, "%I:%M %p")
        except ValueError:
            initObj = datetime.strptime(initT, "%I:%M%p")
            finalObj = datetime.strptime(finalT, "%I:%M%p")
        except:
            ValueError("time format")
    else:
        initObj = datetime.strptime(initT, "%H:%M")
        finalObj = datetime.strptime(finalT, "%H:%M")
    deltaT = finalObj - initObj
    hourDiff = int((deltaT.total_seconds())//(3600))
    return hourDiff


def getMasterFilePath():
    files: list[str] = [file for file in os.listdir('.') if (os.path.isfile(file) and "schedule" in file and "template" in file and file[-5:] == ".xlsx")]
    if len(files) != 1:
        raise FileNotFoundError("Ensure exactly one \"schedule_template.xlsx\" at root dir")
    return files[0]

def getResponsesPath(dirPath: str = "RSP"):
    files = [file for file in os.listdir(dirPath) if (os.path.isfile(os.path.join(dirPath,file)) and file[-5:] == ".xlsx")]
    return files


def analyzeMaster(masterPath: str):
    hourMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    peopleNeededMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    workbook = load_workbook(masterPath)
    sheet = workbook.active
    dayCounter = 0

    for col in range(1, MAX_DAY_IN_WEEK*2 +1, 2):
        timeCounter = 0
        peopleCounter = 0
        isTime: bool = True

        for row in range(2, MAX_SLOT_IN_DAY*2+1):
            val = sheet.cell(row, col).value
            if val is None:
                if(isTime):
                    timeCounter+=1
                else:
                    peopleCounter+=1      
                isTime = not isTime

                continue      

            if(isTime):
                pattern = re.compile("([0-1]?[0-9]|2[0-3])(:)([0-5][0-9])")
                times = [x.group() for x in re.finditer(pattern, val)]
                if(len(times) != 2):
                    raise Exception(f"time format off on {list(Day)[row].name} at slot {col}")
                hourdiff = getTimeDiff(times[0], times[1])
                if hourdiff < 1:
                    hourdiff = 0
                hourMat[dayCounter][timeCounter] = hourdiff
                timeCounter+=1

            else:
                peopleNeededMat[dayCounter][peopleCounter] = int(val)    
                peopleCounter+=1      
            isTime = not isTime

        dayCounter+=1

    print(hourMat)
    print("somehting")
    print(peopleNeededMat)
    writeCSV("PRDAT/shiftHour.csv", hourMat)
    writeCSV("PRDAT/peopleNeeded.csv", peopleNeededMat)
    
def analyzeResponse(filePath: str, writePath: str):
    willMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    workbook = load_workbook(filePath)
    sheet = workbook.active
    dayCounter = 0
    for col in range(2, MAX_DAY_IN_WEEK*2 +1, 2):
        willCounter = 0
        for row in range(2, MAX_SLOT_IN_DAY*2+1, 2):
            val = sheet.cell(row, col).value
            if val is None:
                val = 0
            else:
                try:
                    val = int(val)
                except ValueError:
                    val = 0
                if val < 0:
                    val = 0
                elif val >2:
                    val = 2
            willMat[dayCounter][willCounter] = val
            willCounter +=1
        dayCounter +=1
    
    writeCSV(writePath, willMat)
            


def main():
    #masterPath = getMasterFilePath()
    #analyzeMaster(masterPath)
    files = getResponsesPath()
    csvDir: str = "PRDAT/RSP"
    rspDir: str = "RSP"
    for file in files:
        writePath = os.path.join(csvDir, file[:-5]+".csv")
        readPath = os.path.join(rspDir, file)
        analyzeResponse(readPath, writePath)

main()