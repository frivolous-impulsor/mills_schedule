from openpyxl import load_workbook
import os
import csv
from datetime import datetime
import re
from enum import Enum
import shutil
import subprocess

class Day(Enum):
    SUN = 0
    MON = 1
    TUE = 2
    WED = 3
    THU = 4
    FRI = 5
    SAT = 6


MAX_SLOT_IN_DAY: int = 10
MAX_DAY_IN_WEEK: int = 7

def writeCSV(filePath: str, matrix: list[list[int]]):
    for i, row in enumerate(matrix):
        matrix[i] = [list(Day)[i].name]+row 
    with open(filePath, 'w', newline='') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerows(matrix)

def getTimeDiff(initT: str, finalT: str)->int:
    if("am" in initT.lower() or "pm" in initT.lower() or "am" in finalT.lower() or "pm" in finalT.lower()):
        try: 
            initObj = datetime.strptime(initT, "%I:%M %p")
            finalObj = datetime.strptime(finalT, "%I:%M %p")
        except ValueError:
            initObj = datetime.strptime(initT, "%I:%M%p")
            finalObj = datetime.strptime(finalT, "%I:%M%p")
        except:
            ValueError("time format")
    else:
        initObj = datetime.strptime(initT, "%H:%M")
        finalObj = datetime.strptime(finalT, "%H:%M")
    deltaT = finalObj - initObj
    hourDiff = int((deltaT.total_seconds())//(3600))
    return hourDiff


def getMasterFilePath():
    files: list[str] = [file for file in os.listdir('.') if (os.path.isfile(file) and "schedule" in file and "template" in file and file[-5:] == ".xlsx")]
    if len(files) != 1:
        raise FileNotFoundError("Ensure exactly one \"schedule_template.xlsx\" at root dir")
    return files[0]

def getResponsesPath(dirPath: str = "RSP"):
    files = [file for file in os.listdir(dirPath) if (os.path.isfile(os.path.join(dirPath,file)) and file[-5:] == ".xlsx")]
    return files


def analyzeMaster(masterPath: str):
    hourMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    peopleNeededMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    workbook = load_workbook(masterPath)
    sheet = workbook.active
    dayCounter = 0

    for col in range(1, MAX_DAY_IN_WEEK*2 +1, 2):
        timeCounter = 0
        peopleCounter = 0
        isTime: bool = True

        for row in range(2, MAX_SLOT_IN_DAY*2+1):
            val = sheet.cell(row, col).value
            if val is None:
                if(isTime):
                    timeCounter+=1
                else:
                    peopleCounter+=1      
                isTime = not isTime

                continue      

            if(isTime):
                pattern = re.compile("([0-1]?[0-9]|2[0-3])(:)([0-5][0-9])")
                times = [x.group() for x in re.finditer(pattern, val)]
                if(len(times) != 2):
                    raise Exception(f"time format off on {list(Day)[row].name} at slot {col}")
                hourdiff = getTimeDiff(times[0], times[1])
                if hourdiff < 1:
                    hourdiff = 0
                hourMat[dayCounter][timeCounter] = hourdiff
                timeCounter+=1

            else:
                peopleNeededMat[dayCounter][peopleCounter] = int(val)    
                peopleCounter+=1      
            isTime = not isTime

        dayCounter+=1

    writeCSV("PRDAT/shiftHour.csv", hourMat)
    writeCSV("PRDAT/peopleNeeded.csv", peopleNeededMat)
    
def analyzeResponse(filePath: str, writePath: str):
    willMat = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    workbook = load_workbook(filePath)
    sheet = workbook.active
    dayCounter = 0
    for col in range(2, MAX_DAY_IN_WEEK*2 +1, 2):
        willCounter = 0
        for row in range(2, MAX_SLOT_IN_DAY*2+1, 2):
            val = sheet.cell(row, col).value
            if val is None:
                val = 0
            else:
                try:
                    val = int(val)
                except ValueError:
                    val = 0
                if val < 0:
                    val = 0
                elif val >2:
                    val = 2
            willMat[dayCounter][willCounter] = val
            willCounter +=1
        dayCounter +=1
    
    writeCSV(writePath, willMat)
            


def preconvert():
    masterPath = getMasterFilePath()
    analyzeMaster(masterPath)
    files = getResponsesPath()
    csvDir: str = "PRDAT/RSP"
    rspDir: str = "RSP"
    for file in files:
        writePath = os.path.join(csvDir, file[:-5]+".csv")
        readPath = os.path.join(rspDir, file)
        analyzeResponse(readPath, writePath)

def readCSV(csvPath):
    resultArray: list[list[str]] = [[0 for _ in range(MAX_SLOT_IN_DAY)] for _ in range(MAX_DAY_IN_WEEK)]
    with open(csvPath, mode='r') as csvResultFile:
        resultContent = csv.reader(csvResultFile)
        for i, line in enumerate(resultContent):
            resultArray[i] = line[1:]
    return resultArray


def postconvert():
    template_path: str = "schedule_template.xlsx"
    timeStamp = datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
    resultPath = "result_" + timeStamp + ".xlsx"
    csvSourcePath: str = "PRDAT/result.csv"
    resultDirPath = "RESULT"
    #check and create result dir
    try:
        os.mkdir(resultDirPath)
    except FileExistsError:
        pass
    resultPath = os.path.join(resultDirPath, resultPath)

    shutil.copyfile(template_path, resultPath)
    #read csv
    resultArray = readCSV(csvSourcePath)


    workBook = load_workbook(resultPath)
    sheet = workBook.active
    dayCount = 0
    slotCount = 0
    for col in range(1, MAX_DAY_IN_WEEK*2, 2):
        slotCount = 0
        for row in range(3, MAX_SLOT_IN_DAY*2, 2):
            val: str = resultArray[dayCount][slotCount]
            if(val == ''):
                slotCount +=1
                continue
            val = val.replace(".csv", '')
            sheet.cell(row, col).value = val

            slotCount +=1

        dayCount +=1
    workBook.save(resultPath)
    open('PRDAT/score.txt', 'w').close()


def flushRSP(dirPath: str):
    for fileName in os.listdir(dirPath):
        filePath = os.path.join(dirPath, fileName)
        try:
            if os.path.isfile(filePath) or os.path.islink(filePath):
                os.unlink(filePath)
            elif os.path.isdir(filePath):
                shutil.rmtree(filePath)
        except Exception as e:
            print("fail to delete: %s for %s" % (filePath, e))
        

def main():
    flushRSP("PRDAT/RSP")

    preconvert()
    subprocess.run(["gcc", "SRC/schedule.c", "SRC/indexMaxPriorityQueue.c", "-o", "main"])
    for _ in range(300):
        subprocess.run(["./main"])
    postconvert()


main()